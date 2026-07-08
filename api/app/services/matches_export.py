"""Export a scan's duplicate MATCHES to Excel — one row per record, with record
IDs and deep links into the CRM UI.

Works for ANY scan, including view-only (dry-run) account scans that can't be
merged, so matches can be reviewed, shared, or actioned manually.
"""
import io
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


def record_link(
    crm_type: Optional[str], portal_id: Optional[str], object_type: str, record_id: Optional[str]
) -> str:
    """Deep link to a record in the CRM UI (empty string if not resolvable)."""
    if not record_id:
        return ""
    if crm_type == "salesforce":
        # portal_id is stored as "<orgId>|<instanceUrl>"; the instance URL resolves
        # any 15/18-char record id to its Lightning record view.
        pid = portal_id or ""
        instance = pid.split("|", 1)[1] if "|" in pid else ""
        return f"{instance}/{record_id}" if instance else ""
    if crm_type == "hubspot":
        # HubSpot object type ids: contacts 0-1, companies 0-2, deals 0-3.
        type_id = {"contacts": "0-1", "companies": "0-2", "deals": "0-3"}.get(object_type, "0-1")
        return f"https://app.hubspot.com/contacts/{portal_id}/record/{type_id}/{record_id}"
    return ""


# Display columns per object type: (label, model_key). Name is handled separately.
_FIELDS = {
    "contacts": [("Email", "email"), ("Phone", "phone"), ("Company", "company"), ("Title", "job_title")],
    "companies": [("Domain", "domain"), ("Website", "website"), ("Industry", "industry"), ("Country", "country"), ("Phone", "phone")],
    "accounts": [("Domain", "domain"), ("Website", "website"), ("Industry", "industry"), ("Country", "country"), ("Phone", "phone")],
}


def _display_name(rec: dict, object_type: str) -> str:
    if object_type in ("companies", "accounts"):
        return rec.get("name") or rec.get("company") or ""
    person = " ".join(p for p in [rec.get("first_name"), rec.get("last_name")] if p)
    return person or rec.get("full_name") or rec.get("name") or rec.get("email") or ""


def _status(dup_set: dict) -> str:
    if dup_set.get("merged"):
        return "Merged"
    if dup_set.get("excluded"):
        return "Excluded"
    return (dup_set.get("decision") or "pending").capitalize()


def build_matches_xlsx(scan: dict, connection: dict, duplicate_sets: list) -> bytes:
    """One row per record across all match sets: Set/Confidence/Role/Status/ID/Link
    + object-appropriate fields. Excluded loser records are marked per-record."""
    object_type = scan.get("object_type") or "contacts"
    crm_type = connection.get("crm_type")
    portal_id = connection.get("portal_id")

    fields = _FIELDS.get(object_type, _FIELDS["contacts"])
    headers = ["Set", "Confidence", "Role", "Status", "Record ID", "Link", "Name"] + [
        label for label, _ in fields
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Matches"
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="642585")
        cell.alignment = Alignment(horizontal="left")
    ws.freeze_panes = "A2"

    for i, s in enumerate(duplicate_sets, 1):
        conf = round(float(s.get("confidence") or 0), 1)
        status = _status(s)
        excluded_ids = set(s.get("excluded_record_ids") or [])
        winner = s.get("winner_data") or {}
        losers = s.get("loser_data") or []

        rows = [("Kept", winner)] + [("Duplicate", l or {}) for l in losers]
        for role, rec in rows:
            rid = rec.get("id") or ""
            # A loser the reviewer marked "not a duplicate" is flagged, not merged.
            role_label = role
            if role == "Duplicate" and rid in excluded_ids:
                role_label = "Not a duplicate"
            link = record_link(crm_type, portal_id, object_type, rid)
            row = [i, conf, role_label, status, rid, link, _display_name(rec, object_type)]
            row += [rec.get(key) for _, key in fields]
            ws.append(row)
            if link:
                c = ws.cell(row=ws.max_row, column=6)
                c.hyperlink = link
                c.font = Font(color="0563C1", underline="single")

    for col in ws.columns:
        longest = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max(longest + 2, 10), 60)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
