"""Report generation service."""
from datetime import datetime, timezone
from typing import Optional
import io

from app.services.supabase_client import get_supabase
from app.services.tenancy import can_access_tenant

# WeasyPrint requires GTK system libraries (gobject, pango).
# Import lazily so the app starts without them; PDF generation
# will fail with a clear error if the libs are missing.
_weasyprint = None

def _get_weasyprint():
    global _weasyprint
    if _weasyprint is None:
        try:
            import weasyprint as wp
            _weasyprint = wp
        except OSError as e:
            raise RuntimeError(
                "WeasyPrint requires GTK system libraries. "
                "Install them (e.g., brew install pango on macOS) or use Docker."
            ) from e
    return _weasyprint


class ReportService:
    """Service for generating deduplication reports."""

    def __init__(self):
        self.supabase = get_supabase()

    async def generate_report(
        self,
        merge_id: str,
        user_id: str,
    ) -> dict:
        """
        Generate a report for a completed merge.

        Args:
            merge_id: The merge ID
            user_id: The user ID

        Returns:
            Dict with report data
        """
        # Get merge details — scoped to the caller's tenant access (defense in
        # depth; the router also checks access). The service-role client bypasses
        # RLS, so the tenant check is explicit.
        merge_result = self.supabase.table("merges").select("*").eq(
            "id", merge_id
        ).limit(1).execute()
        merge = (merge_result.data or [None])[0]

        if not merge or not can_access_tenant(
            self.supabase, merge.get("tenant_id"), user_id
        ):
            raise Exception("Merge not found")

        # Get scan details
        scan_result = self.supabase.table("scans").select("*").eq(
            "id", merge["scan_id"]
        ).single().execute()

        scan = scan_result.data or {}

        # Get connection details
        conn_result = self.supabase.table("crm_connections").select("*").eq(
            "id", scan.get("connection_id")
        ).single().execute()

        connection = conn_result.data or {}

        # Build report data
        report_data = {
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "crm_type": connection.get("crm_type", "unknown"),
            "portal_id": connection.get("portal_id"),
            "scan": {
                "id": scan.get("id"),
                "object_type": scan.get("object_type"),
                "records_scanned": scan.get("records_scanned", 0),
                "duplicates_found": scan.get("duplicates_found", 0),
                "started_at": scan.get("started_at"),
                "completed_at": scan.get("completed_at"),
            },
            "merge": {
                "id": merge["id"],
                "total_sets": merge["total_sets"],
                "completed_sets": merge["completed_sets"],
                "failed_sets": merge["failed_sets"],
                "success_rate": round(
                    (merge["completed_sets"] / max(merge["total_sets"], 1)) * 100, 1
                ),
                "started_at": merge.get("started_at"),
                "completed_at": merge.get("completed_at"),
            },
            "summary": {
                "records_removed": merge["completed_sets"],  # Each merge removes 1+ records
                "data_quality_improvement": f"{round((merge['completed_sets'] / max(scan.get('records_scanned', 1), 1)) * 100, 1)}%",
            },
        }

        # Per-record detail: for each merged set, the surviving record + the records
        # merged into it, with ALL captured fields (from the pre-merge backups).
        # Columns depend on the object type — companies have no person fields.
        if scan.get("object_type") == "companies":
            report_fields = [
                {"key": "name", "label": "Company Name"},
                {"key": "domain", "label": "Domain"},
                {"key": "website", "label": "Website"},
                {"key": "phone", "label": "Phone"},
                {"key": "industry", "label": "Industry"},
                {"key": "country", "label": "Country"},
                {"key": "created_at", "label": "Created"},
                {"key": "updated_at", "label": "Updated"},
                {"key": "association_count", "label": "Associations"},
            ]
        else:
            report_fields = [
                {"key": "first_name", "label": "First Name"},
                {"key": "last_name", "label": "Last Name"},
                {"key": "email", "label": "Email"},
                {"key": "phone", "label": "Phone"},
                {"key": "company", "label": "Company"},
                {"key": "job_title", "label": "Job Title"},
                {"key": "created_at", "label": "Created"},
                {"key": "updated_at", "label": "Updated"},
                {"key": "association_count", "label": "Associations"},
            ]

        def _record(c: dict, role: str) -> dict:
            c = c or {}
            rec = {"role": role, "id": c.get("id")}
            for f in report_fields:
                rec[f["key"]] = c.get(f["key"])
            return rec

        backups = self.supabase.table("merge_backups").select(
            "winner_snapshot,loser_record_ids,loser_snapshot"
        ).eq("merge_id", merge_id).execute()

        merged_sets = []
        for b in (backups.data or []):
            snaps = b.get("loser_snapshot") or []
            kept_ids = set(b.get("loser_record_ids") or [])
            losers = [l for l in snaps if (not kept_ids) or (l or {}).get("id") in kept_ids]
            records = [_record(b.get("winner_snapshot") or {}, "Surviving")]
            records += [_record(l, "Merged") for l in losers]
            merged_sets.append({"records": records})
        report_data["merged_sets"] = merged_sets
        report_data["report_fields"] = report_fields

        # Save report (stamped with the merge's tenant; user_id is the actor).
        report_id = await self._save_report(
            merge_id, user_id, merge["tenant_id"], report_data
        )
        report_data["id"] = report_id

        return report_data

    async def _save_report(
        self,
        merge_id: str,
        user_id: str,
        tenant_id: str,
        report_data: dict,
    ) -> str:
        """Save report to database."""
        import uuid
        report_id = str(uuid.uuid4())

        self.supabase.table("reports").insert({
            "id": report_id,
            "merge_id": merge_id,
            "tenant_id": tenant_id,
            "user_id": user_id,
            "report_data": report_data,
        }).execute()

        return report_id

    async def generate_pdf(self, report_id: str, user_id: str) -> bytes:
        """
        Generate PDF from report data.

        Args:
            report_id: The report ID
            user_id: The owner (report is scoped to this user — defense in depth)

        Returns:
            PDF bytes
        """
        # Get report — scoped to tenant access so the ID alone can't leak another
        # tenant's report even if a future caller skips the router check.
        result = self.supabase.table("reports").select("*").eq(
            "id", report_id
        ).limit(1).execute()
        row = (result.data or [None])[0]

        if not row or not can_access_tenant(
            self.supabase, row.get("tenant_id"), user_id
        ):
            raise Exception("Report not found")

        report = row["report_data"]

        # Generate HTML
        html_content = self._generate_html(report)

        # Convert to PDF
        wp = _get_weasyprint()
        pdf_bytes = wp.HTML(string=html_content).write_pdf(
            stylesheets=[wp.CSS(string=self._get_pdf_styles())]
        )

        return pdf_bytes

    async def generate_html(self, report_id: str, user_id: str) -> str:
        """Standalone HTML report — no native libraries required (weasyprint-free).
        Same content as the PDF, with the report CSS inlined so it renders on its
        own; viewable in the browser and printable to PDF from there."""
        result = self.supabase.table("reports").select("*").eq(
            "id", report_id
        ).limit(1).execute()
        row = (result.data or [None])[0]

        if not row or not can_access_tenant(
            self.supabase, row.get("tenant_id"), user_id
        ):
            raise Exception("Report not found")

        report = row["report_data"]
        html = self._generate_html(report)
        style = f"<style>{self._get_pdf_styles()}</style>"
        if "</head>" in html:
            return html.replace("</head>", style + "</head>", 1)
        if "<body" in html:
            return html.replace("<body", style + "<body", 1)
        return style + html

    async def generate_xlsx(self, report_id: str, user_id: str) -> bytes:
        """Excel workbook of the merged-record detail — one row per record, every
        captured field as a column. Pure-Python (openpyxl), no native libraries."""
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        result = self.supabase.table("reports").select("*").eq(
            "id", report_id
        ).limit(1).execute()
        row = (result.data or [None])[0]
        if not row or not can_access_tenant(
            self.supabase, row.get("tenant_id"), user_id
        ):
            raise Exception("Report not found")

        report = row["report_data"]
        merged_sets = report.get("merged_sets", []) or []
        fields = report.get("report_fields") or []

        wb = Workbook()
        ws = wb.active
        ws.title = "Merged Records"
        ws.append(["Set", "Role", "Record ID"] + [f["label"] for f in fields])
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="642585")
            cell.alignment = Alignment(horizontal="left")
        ws.freeze_panes = "A2"

        for i, s in enumerate(merged_sets, 1):
            for r in (s.get("records") or []):
                ws.append([i, r.get("role"), r.get("id")] + [r.get(f["key"]) for f in fields])

        for col in ws.columns:
            longest = max((len(str(c.value)) for c in col if c.value is not None), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(max(longest + 2, 10), 42)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def _generate_html(self, report: dict) -> str:
        """Generate HTML for PDF."""
        scan = report.get("scan", {})
        merge = report.get("merge", {})
        summary = report.get("summary", {})
        merged_sets = report.get("merged_sets", []) or []
        report_fields = report.get("report_fields") or []

        def _esc(v: object) -> str:
            return str(v if v not in (None, "") else "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")

        def _rname(r: dict) -> str:
            # Contacts show person name/email; companies fall back to company name.
            return (
                " ".join(p for p in [r.get("first_name"), r.get("last_name")] if p)
                or r.get("name") or r.get("email") or "(record)"
            )

        set_blocks = []
        for i, s in enumerate(merged_sets, 1):
            records = s.get("records") or []
            if not records:
                continue
            # Only show fields that are populated on at least one record in this set.
            fields = [f for f in report_fields if any(r.get(f["key"]) not in (None, "") for r in records)]
            surv = "background:#eafaf0"
            ths = "".join(
                f"<th style='text-align:left;padding:6px;border-bottom:2px solid #ccc;{surv if r.get('role')=='Surviving' else ''}'>"
                f"{'&#10003; ' if r.get('role')=='Surviving' else ''}{_esc(_rname(r))}"
                f"<br><span style='font-weight:normal;color:#888;font-size:11px'>{_esc(r.get('role'))}</span></th>"
                for r in records
            )
            body = ""
            for f in fields:
                tds = "".join(
                    f"<td style='padding:6px;border-bottom:1px solid #eee;{'background:#f4fbf7' if r.get('role')=='Surviving' else ''}'>{_esc(r.get(f['key']))}</td>"
                    for r in records
                )
                body += (
                    f"<tr><td style='padding:6px;border-bottom:1px solid #eee;color:#666;font-weight:600'>{_esc(f['label'])}</td>{tds}</tr>"
                )
            set_blocks.append(
                f"<h3 style='margin:18px 0 6px'>Set {i}</h3>"
                "<table style='width:100%;border-collapse:collapse;font-size:12px'>"
                "<thead><tr><th style='text-align:left;padding:6px;border-bottom:2px solid #ccc'>Field</th>"
                f"{ths}</tr></thead><tbody>{body}</tbody></table>"
            )

        merged_section = ""
        if set_blocks:
            merged_section = (
                f"<div class='section'><h2>Merged Records ({len(merged_sets)} sets)</h2>"
                f"{''.join(set_blocks)}</div>"
            )

        return f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="utf-8">
            <title>CRM Deduplication Report</title>
        </head>
        <body>
            <div class="header">
                <h1>CRM Deduplication Report</h1>
                <p class="subtitle">Generated: {self._format_date(report.get('generated_at'))}</p>
            </div>

            <div class="section">
                <h2>Overview</h2>
                <table class="info-table">
                    <tr>
                        <td class="label">CRM Platform</td>
                        <td>{report.get('crm_type', 'N/A').title()}</td>
                    </tr>
                    <tr>
                        <td class="label">Portal ID</td>
                        <td>{report.get('portal_id', 'N/A')}</td>
                    </tr>
                    <tr>
                        <td class="label">Object Type</td>
                        <td>{scan.get('object_type', 'N/A').title()}</td>
                    </tr>
                </table>
            </div>

            <div class="section">
                <h2>Scan Results</h2>
                <div class="stats-grid">
                    <div class="stat-box">
                        <div class="stat-value">{scan.get('records_scanned', 0):,}</div>
                        <div class="stat-label">Records Scanned</div>
                    </div>
                    <div class="stat-box">
                        <div class="stat-value">{scan.get('duplicates_found', 0):,}</div>
                        <div class="stat-label">Duplicate Sets Found</div>
                    </div>
                </div>
            </div>

            <div class="section">
                <h2>Merge Results</h2>
                <div class="stats-grid">
                    <div class="stat-box success">
                        <div class="stat-value">{merge.get('completed_sets', 0):,}</div>
                        <div class="stat-label">Successfully Merged</div>
                    </div>
                    <div class="stat-box">
                        <div class="stat-value">{merge.get('failed_sets', 0):,}</div>
                        <div class="stat-label">Failed</div>
                    </div>
                    <div class="stat-box">
                        <div class="stat-value">{merge.get('success_rate', 0)}%</div>
                        <div class="stat-label">Success Rate</div>
                    </div>
                </div>
            </div>

            <div class="section">
                <h2>Impact Summary</h2>
                <table class="info-table">
                    <tr>
                        <td class="label">Duplicate Records Removed</td>
                        <td><strong>{summary.get('records_removed', 0):,}</strong></td>
                    </tr>
                    <tr>
                        <td class="label">Data Quality Improvement</td>
                        <td><strong>{summary.get('data_quality_improvement', 'N/A')}</strong></td>
                    </tr>
                </table>
            </div>

            {merged_section}

            <div class="footer">
                <p>Report generated by CRM Dedup Tool</p>
                <p>Powered by LeanScale</p>
            </div>
        </body>
        </html>
        """

    def _get_pdf_styles(self) -> str:
        """Get CSS styles for PDF."""
        return """
        @page {
            size: letter;
            margin: 1in;
        }

        body {
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
            font-size: 12pt;
            color: #333;
            line-height: 1.5;
        }

        .header {
            text-align: center;
            margin-bottom: 30px;
            border-bottom: 2px solid #2563eb;
            padding-bottom: 20px;
        }

        h1 {
            color: #2563eb;
            margin: 0;
            font-size: 24pt;
        }

        .subtitle {
            color: #666;
            margin: 5px 0 0 0;
        }

        .section {
            margin-bottom: 25px;
        }

        h2 {
            color: #1f2937;
            font-size: 14pt;
            border-bottom: 1px solid #e5e7eb;
            padding-bottom: 5px;
            margin-bottom: 15px;
        }

        .info-table {
            width: 100%;
            border-collapse: collapse;
        }

        .info-table td {
            padding: 8px 0;
            border-bottom: 1px solid #f3f4f6;
        }

        .info-table .label {
            color: #6b7280;
            width: 200px;
        }

        .stats-grid {
            display: flex;
            gap: 20px;
        }

        .stat-box {
            background: #f9fafb;
            padding: 15px 20px;
            border-radius: 8px;
            flex: 1;
            text-align: center;
        }

        .stat-box.success {
            background: #ecfdf5;
        }

        .stat-value {
            font-size: 24pt;
            font-weight: bold;
            color: #1f2937;
        }

        .stat-box.success .stat-value {
            color: #059669;
        }

        .stat-label {
            font-size: 10pt;
            color: #6b7280;
            margin-top: 5px;
        }

        .footer {
            margin-top: 40px;
            text-align: center;
            color: #9ca3af;
            font-size: 10pt;
            border-top: 1px solid #e5e7eb;
            padding-top: 20px;
        }

        .footer p {
            margin: 2px 0;
        }
        """

    def _format_date(self, date_str: Optional[str]) -> str:
        """Format ISO date string for display."""
        if not date_str:
            return "N/A"
        try:
            dt = datetime.fromisoformat(date_str.replace("Z", "+00:00"))
            return dt.strftime("%B %d, %Y at %I:%M %p")
        except (ValueError, TypeError):
            return date_str
